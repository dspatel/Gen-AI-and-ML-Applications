\
"""
Rolling 30-Minute Range Breakout Monitor (TradingView candles via tvDatafeed) + Excel log (+ optional email)

Implements your requested logic:
1) Build a range using the NEXT 30 minutes of candles (for 15m interval => 2 candles)
2) After range is ready, wait for breakout (default breakout reference = CLOSE after candle closes)
3) When breakout occurs, log the breakout candle + indicators + targets
4) Immediately start building the NEXT 30-minute range window from the breakout onward
5) Repeat all day

Targets:
- Range-based targets with an ATR floor:
    W = range_high - range_low
    step = max(W * TARGET_PCT, ATR_FLOOR_MULT * ATR14)
    Up targets:  range_high + k*step
    Down targets: range_low  - k*step
"""

import json
import time
import smtplib
import logging
from dataclasses import dataclass
from email.mime.text import MIMEText
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import keyring
from tvDatafeed import TvDatafeed, Interval

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# =========================
# USER SETTINGS
# =========================
SYMBOLS_FILE = "symbols.json"

CANDLE_INTERVAL = Interval.in_15_minute

OR_MINUTES = 30
BREAKOUT_REF = "close"      # "close" or "highlow"

TARGET_PCT = 1.0
NUM_TARGETS = 3
USE_ATR_FLOOR = True
ATR_FLOOR_MULT = 0.5

POLL_SECONDS = 30
HISTORY_BARS = 400
EXCEL_LOG_PATH = "rolling_range_breakouts.xlsx"

SEND_EMAIL_ALERTS = False


# Logging toggles
LOG_RANGE_READY_TO_EXCEL = True   # set False to log ONLY breakouts
LOG_BREAKOUT_TO_EXCEL = True      # set False to disable breakout rows (not recommended)

# Notification toggles
EMAIL_ON_BREAKOUT = True          # if SEND_EMAIL_ALERTS=True, email only when this is True

# Session / shutdown behavior
EXIT_AT_MARKET_CLOSE = True           # stop script when regular session ends
WRITE_DAILY_SUMMARY_TO_EXCEL = True   # append summary to Excel at exit

# Testing / simulation (for after-hours testing)
TEST_MODE = False                     # True => replay recent historical bars quickly (no waiting)
TEST_BARS_TO_REPLAY = 400             # how many most-recent closed bars to simulate per symbol
FORCE_DISABLE_EMAIL_IN_TEST = True
SEPARATE_TEST_LOGS = True            # if True, TEST_MODE writes to a separate Excel file
TEST_LOG_FILE_PREFIX = "rolling_range_breakouts_TEST"
DEBUG = True

MARKET_HOURS_ONLY = True
RUN_EXTENDED_HOURS = False
US_MARKET_TZ = "America/New_York"
MARKET_OPEN = (9, 30)
MARKET_CLOSE = (16, 0)

# Indicator settings for logging
ATR_LEN = 14
EMA_FAST = 20
EMA_MID = 50
EMA_SLOW = 200
RSI_LEN = 14
VOL_SMA_LEN = 20


# =========================
# EMAIL (optional)
# =========================
@dataclass
class EmailConfig:
    smtp_host: str
    smtp_port: int
    username: str
    password: str
    from_addr: str
    to_addrs: List[str]
    use_tls: bool = True


EMAIL_CFG = EmailConfig(
    smtp_host="smtp.office365.com",
    smtp_port=587,
    username="YOUR_OUTLOOK_EMAIL@domain.com",
    password="YOUR_PASSWORD_OR_APP_PASSWORD",
    from_addr="YOUR_OUTLOOK_EMAIL@domain.com",
    to_addrs=["YOUR_OUTLOOK_EMAIL@domain.com"],
    use_tls=True
)


# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=logging.DEBUG if DEBUG else logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
log = logging.getLogger("ROLLING_ORB")



# =========================
# DAILY STATS (for end-of-day snapshot)
# =========================
@dataclass
class SymbolDailyStats:
    symbol: str
    exchange_used: str = ""
    ranges_built: int = 0

    breakouts_total: int = 0
    breakouts_up: int = 0
    breakouts_down: int = 0

    sum_distance: float = 0.0
    max_distance: float = 0.0

    last_event_time: str = ""
    last_breakout_close: float = float("nan")

    def record_range_ready(self, event_time: str) -> None:
        self.ranges_built += 1
        self.last_event_time = event_time or self.last_event_time

    def record_breakout(self, direction: str, event_time: str, distance: float, close: float) -> None:
        self.breakouts_total += 1
        if direction == "UP":
            self.breakouts_up += 1
        else:
            self.breakouts_down += 1

        d = float(distance or 0.0)
        self.sum_distance += d
        self.max_distance = max(self.max_distance, d)

        self.last_event_time = event_time or self.last_event_time
        self.last_breakout_close = close

    def avg_distance(self) -> float:
        return (self.sum_distance / self.breakouts_total) if self.breakouts_total else 0.0


def print_daily_summary(stats_map: Dict[str, SymbolDailyStats]) -> None:
    log.info("========== DAILY SNAPSHOT ==========")
    for sym, st in stats_map.items():
        log.info(
            f"{sym} | ranges={st.ranges_built} | breakouts={st.breakouts_total} "
            f"(UP={st.breakouts_up}, DOWN={st.breakouts_down}) | "
            f"avg_dist={st.avg_distance():.4f} | max_dist={st.max_distance:.4f} | "
            f"last_time={st.last_event_time} | last_breakout_close={st.last_breakout_close}"
        )
    log.info("====================================")


def write_daily_summary(excel_path: str, stats_map: Dict[str, SymbolDailyStats]) -> None:
    """Append one summary row per symbol to a sheet named 'daily_summary'."""
    from openpyxl import load_workbook
    from datetime import datetime

    wb = load_workbook(excel_path)

    if "daily_summary" not in wb.sheetnames:
        ws = wb.create_sheet("daily_summary")
        ws.append([
            "run_end_time_local",
            "symbol",
            "exchange_used",
            "ranges_built",
            "breakouts_total",
            "breakouts_up",
            "breakouts_down",
            "avg_distance_past_range",
            "max_distance_past_range",
            "last_event_time",
            "last_breakout_close",
        ])
    else:
        ws = wb["daily_summary"]

    end_time = datetime.now().isoformat(timespec="seconds")

    for sym, st in stats_map.items():
        ws.append([
            end_time,
            sym,
            st.exchange_used,
            st.ranges_built,
            st.breakouts_total,
            st.breakouts_up,
            st.breakouts_down,
            st.avg_distance(),
            st.max_distance,
            st.last_event_time,
            None if np.isnan(st.last_breakout_close) else st.last_breakout_close,
        ])

    wb.save(excel_path)

# =========================
# MARKET HOURS CHECK
# =========================
def in_us_market_hours() -> bool:
    from datetime import datetime, time as dtime
    import zoneinfo
    tz = zoneinfo.ZoneInfo(US_MARKET_TZ)
    now = datetime.now(tz)

    if now.weekday() >= 5:
        return False

    if RUN_EXTENDED_HOURS:
        return True

    open_t = dtime(MARKET_OPEN[0], MARKET_OPEN[1])
    close_t = dtime(MARKET_CLOSE[0], MARKET_CLOSE[1])
    return open_t <= now.time() <= close_t


# =========================
# EMAIL
# =========================
def send_email(cfg: EmailConfig, subject: str, body: str) -> None:
    msg = MIMEText(body, "plain")
    msg["Subject"] = subject
    msg["From"] = cfg.from_addr
    msg["To"] = ", ".join(cfg.to_addrs)

    with smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=30) as server:
        if cfg.use_tls:
            server.starttls()
        server.login(cfg.username, cfg.password)
        server.sendmail(cfg.from_addr, cfg.to_addrs, msg.as_string())


# =========================
# INDICATORS
# =========================
def sma(series: pd.Series, length: int) -> pd.Series:
    return series.rolling(length, min_periods=length).mean()

def ema(series: pd.Series, length: int) -> pd.Series:
    return series.ewm(span=length, adjust=False, min_periods=length).mean()

def atr_wilder(df: pd.DataFrame, length: int = 14) -> pd.Series:
    high = df["high"]
    low = df["low"]
    close = df["close"]
    prev_close = close.shift(1)
    tr = pd.concat([(high - low), (high - prev_close).abs(), (low - prev_close).abs()], axis=1).max(axis=1)
    return tr.ewm(alpha=1/length, adjust=False, min_periods=length).mean()

def rsi(series: pd.Series, length: int = 14) -> pd.Series:
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1/length, adjust=False, min_periods=length).mean()
    avg_loss = loss.ewm(alpha=1/length, adjust=False, min_periods=length).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))

def crossover(prev_val: float, curr_val: float, level: float) -> bool:
    return (prev_val <= level) and (curr_val > level)

def crossunder(prev_val: float, curr_val: float, level: float) -> bool:
    return (prev_val >= level) and (curr_val < level)


# =========================
# TARGETS (Range + ATR floor)
# =========================
def compute_targets(range_high: float, range_low: float, atr14: Optional[float]) -> dict:
    W = float(range_high - range_low)
    step = W * TARGET_PCT
    if USE_ATR_FLOOR and atr14 is not None and not np.isnan(atr14):
        step = max(step, ATR_FLOOR_MULT * float(atr14))

    up = [range_high + i * step for i in range(1, NUM_TARGETS + 1)]
    dn = [range_low  - i * step for i in range(1, NUM_TARGETS + 1)]

    return {
        "range_width": W,
        "target_step": float(step),
        "up_targets": up,
        "dn_targets": dn,
        "next_high_target": up[0] if up else None,
        "next_low_target": dn[0] if dn else None
    }


# =========================
# SYMBOL CONFIG + FETCH
# =========================
def load_symbols_config(path: str) -> Tuple[List[dict], List[str]]:
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    fallback = cfg.get("default_exchange_fallback", ["AMEX", "NASDAQ", "NYSE"])
    symbols = [s for s in cfg.get("symbols", []) if s.get("enabled", True)]
    return symbols, fallback

def normalize_tv_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().reset_index()
    if "datetime" in df.columns:
        df.rename(columns={"datetime": "time"}, inplace=True)
    elif "date" in df.columns:
        df.rename(columns={"date": "time"}, inplace=True)
    else:
        df.rename(columns={df.columns[0]: "time"}, inplace=True)

    needed = ["time", "open", "high", "low", "close", "volume"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    return df[needed].sort_values("time").reset_index(drop=True)

def fetch_with_exchange_fallback(
    tv: TvDatafeed,
    symbol: str,
    preferred_exchange: Optional[str],
    fallback_exchanges: List[str],
    interval: Interval,
    n_bars: int,
    retries_per_exchange: int = 2
) -> Tuple[pd.DataFrame, Optional[str]]:
    exchanges_to_try: List[str] = []
    if preferred_exchange:
        exchanges_to_try.append(preferred_exchange)
    for ex in fallback_exchanges:
        if ex not in exchanges_to_try:
            exchanges_to_try.append(ex)

    last_err = None
    for exch in exchanges_to_try:
        for attempt in range(1, retries_per_exchange + 1):
            try:
                raw = tv.get_hist(symbol=symbol, exchange=exch, interval=interval, n_bars=n_bars)
                if raw is None or raw.empty:
                    raise ValueError("Empty dataframe")
                return normalize_tv_df(raw), exch
            except Exception as e:
                last_err = e
                time.sleep(0.5 * attempt)

    log.warning(f"[{symbol}] no data from exchanges {exchanges_to_try}. last_err={last_err}")
    return pd.DataFrame(), None


# =========================
# RANGE STATE
# =========================
def interval_to_minutes(interval: Interval) -> int:
    mapping = {
        Interval.in_1_minute: 1,
        Interval.in_3_minute: 3,
        Interval.in_5_minute: 5,
        Interval.in_15_minute: 15,
        Interval.in_30_minute: 30,
        Interval.in_45_minute: 45,
        Interval.in_1_hour: 60,
        Interval.in_2_hour: 120,
        Interval.in_3_hour: 180,
        Interval.in_4_hour: 240,
        Interval.in_daily: 1440,
        Interval.in_weekly: 10080,
    }
    if interval not in mapping:
        raise ValueError(f"Unsupported interval for OR_MINUTES calc: {interval}")
    return mapping[interval]

@dataclass
class RollingRangeState:
    phase: str = "BUILD_RANGE"
    bars_needed: int = 0
    bars_collected: int = 0

    range_high: float = np.nan
    range_low: float = np.nan
    range_start_time: Optional[str] = None
    range_end_time: Optional[str] = None

    range_width: Optional[float] = None
    target_step: Optional[float] = None
    up_targets: Optional[List[float]] = None
    dn_targets: Optional[List[float]] = None

    last_bar_time_seen: Optional[str] = None


# =========================
# EXCEL LOGGING
# =========================
EXCEL_HEADERS = [
    "log_time_local","symbol","exchange_used","interval","event_type","event_time",
    "range_start","range_end","range_high","range_low","range_width","target_step",
    "breakout_dir","breakout_ref","distance_past_range",
    "open","high","low","close","volume",
    "ema20","ema50","ema200","atr14","rsi14","vol_sma20",
    "up_T1","up_T2","up_T3","dn_T1","dn_T2","dn_T3"
]

def ensure_excel_exists(path: str) -> None:
    try:
        load_workbook(path)
        return
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "events"
        ws.append(EXCEL_HEADERS)
        for i, h in enumerate(EXCEL_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 14), 45)
        wb.save(path)

def append_excel_row(path: str, row: List) -> None:
    wb = load_workbook(path)
    ws = wb["events"]
    ws.append(row)
    wb.save(path)


# =========================
# CORE EVALUATION
# =========================
def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["ema20"] = ema(df["close"], EMA_FAST)
    df["ema50"] = ema(df["close"], EMA_MID)
    df["ema200"] = ema(df["close"], EMA_SLOW)
    df["atr14"] = atr_wilder(df, ATR_LEN)
    df["rsi14"] = rsi(df["close"], RSI_LEN)
    df["vol_sma20"] = sma(df["volume"], VOL_SMA_LEN)
    return df

def get_ref_values(row: pd.Series, ref: str) -> Tuple[float, float]:
    if ref == "highlow":
        return float(row["high"]), float(row["low"])
    return float(row["close"]), float(row["close"])

def evaluate_rolling_range(symbol: str, df: pd.DataFrame, state: RollingRangeState, interval: Interval) -> List[dict]:
    events: List[dict] = []
    if len(df) < 5:
        return events

    df = add_indicators(df)
    curr = df.iloc[-1]
    prev = df.iloc[-2]

    if state.bars_needed == 0:
        mins = interval_to_minutes(interval)
        state.bars_needed = int(np.ceil(OR_MINUTES / mins)) or 1

    # Phase A: Build range window (next 30 minutes)
    if state.phase == "BUILD_RANGE":
        if state.bars_collected == 0:
            state.range_high = float(curr["high"])
            state.range_low = float(curr["low"])
            state.range_start_time = str(curr["time"])
        else:
            state.range_high = max(state.range_high, float(curr["high"]))
            state.range_low = min(state.range_low, float(curr["low"]))

        state.bars_collected += 1
        state.range_end_time = str(curr["time"])

        if state.bars_collected >= state.bars_needed:
            atr14 = float(curr["atr14"]) if not np.isnan(curr["atr14"]) else None
            t = compute_targets(state.range_high, state.range_low, atr14)
            state.range_width = t["range_width"]
            state.target_step = t["target_step"]
            state.up_targets = t["up_targets"]
            state.dn_targets = t["dn_targets"]
            state.next_high_target = t.get("next_high_target")
            state.next_low_target = t.get("next_low_target")

            state.phase = "WAIT_BREAKOUT"
            events.append({
                "type": "RANGE_READY",
                "time": str(curr["time"]),
                "range_start": state.range_start_time,
                "range_end": state.range_end_time,
                "range_high": state.range_high,
                "range_low": state.range_low,
                "range_width": state.range_width,
                "target_step": state.target_step,
                "up_targets": state.up_targets,
                "dn_targets": state.dn_targets,
                "row": curr.to_dict()
            })
        return events

    # Phase B: Wait for breakout
    prev_up, prev_dn = get_ref_values(prev, BREAKOUT_REF)
    curr_up, curr_dn = get_ref_values(curr, BREAKOUT_REF)

    breakout_up = crossover(prev_up, curr_up, state.range_high)
    breakout_dn = crossunder(prev_dn, curr_dn, state.range_low)

    if breakout_up or breakout_dn:
        direction = "UP" if breakout_up else "DOWN"
        close = float(curr["close"])
        dist = (close - state.range_high) if direction == "UP" else (state.range_low - close)

        events.append({
            "type": "BREAKOUT",
            "direction": direction,
            "time": str(curr["time"]),
            "range_start": state.range_start_time,
            "range_end": state.range_end_time,
            "range_high": state.range_high,
            "range_low": state.range_low,
            "range_width": state.range_width,
            "target_step": state.target_step,
            "up_targets": state.up_targets,
            "dn_targets": state.dn_targets,
            "breakout_ref": BREAKOUT_REF,
            "distance_past_range": float(dist),
            "row": curr.to_dict()
        })

        # Reset: build next 30-min range from next candles after breakout
        state.phase = "BUILD_RANGE"
        state.bars_collected = 0
        state.range_high = np.nan
        state.range_low = np.nan
        state.range_start_time = None
        state.range_end_time = None
        state.range_width = None
        state.target_step = None
        state.up_targets = None
        state.dn_targets = None
        state.next_high_target = None
        state.next_low_target = None

    return events


# =========================
# TV credentials
# =========================
def load_tv_credentials_from_keyring() -> Tuple[Optional[str], Optional[str]]:
    user = keyring.get_password("tradingview", "username")
    pwd = keyring.get_password("tradingview", "password")
    return user, pwd


# =========================
# MAIN LOOP
# =========================


def run_test_mode(tv: TvDatafeed, symbols_cfg: List[dict], fallback_exchanges: List[str],
                  states: Dict[str, RollingRangeState], stats: Dict[str, SymbolDailyStats]) -> None:
    """
    After-hours testing mode:
    - Fetch recent historical bars for each symbol
    - Replay the last TEST_BARS_TO_REPLAY bars quickly (no sleeping)
    - Generates RANGE_READY/BREAKOUT events and writes to Excel (respecting Excel toggles)
    - Prints + writes daily summary and exits
    """
    log.warning("TEST_MODE=ON: replaying historical bars quickly. No waiting.")
    for s in symbols_cfg:
        sym = s["symbol"]
        preferred_exch = s.get("exchange")

        df, exch_used = fetch_with_exchange_fallback(
            tv, sym, preferred_exch, fallback_exchanges,
            CANDLE_INTERVAL, max(HISTORY_BARS, TEST_BARS_TO_REPLAY + 50)
        )
        if df.empty or len(df) < 10:
            continue
        if exch_used:
            stats[sym].exchange_used = exch_used

        start = max(5, len(df) - TEST_BARS_TO_REPLAY)
        for i in range(start, len(df)):
            view = df.iloc[: i + 1].copy()
            bar_time = str(view.iloc[-1]["time"])
            if states[sym].last_bar_time_seen == bar_time:
                continue
            states[sym].last_bar_time_seen = bar_time

            events = evaluate_rolling_range(sym, view, states[sym], CANDLE_INTERVAL)
            for ev in events:
                event_type = ev.get("type")

                if event_type == "RANGE_READY" and not LOG_RANGE_READY_TO_EXCEL:
                    continue
                if event_type == "BREAKOUT" and not LOG_BREAKOUT_TO_EXCEL:
                    continue

                row = ev.get("row", {})

                if event_type == "RANGE_READY":
                    stats[sym].record_range_ready(ev.get("time", ""))
                elif event_type == "BREAKOUT":
                    stats[sym].record_breakout(
                        direction=ev.get("direction", ""),
                        event_time=ev.get("time", ""),
                        distance=ev.get("distance_past_range", 0.0),
                        close=float(row.get("close")) if row.get("close") is not None else float("nan")
                    )

                up = ev.get("up_targets") or []
                dn = ev.get("dn_targets") or []
                up_pad = (up + [None]*3)[:3]
                dn_pad = (dn + [None]*3)[:3]

                from datetime import datetime
                xl = [
                    datetime.now().isoformat(timespec="seconds"),
                    sym,
                    exch_used,
                    str(CANDLE_INTERVAL),
                    event_type,
                    ev.get("time"),
                    ev.get("range_start"),
                    ev.get("range_end"),
                    ev.get("range_high"),
                    ev.get("range_low"),
                    ev.get("range_width"),
                    ev.get("next_high_target"),
                    ev.get("next_low_target"),
                    ev.get("target_step"),
                    up_pad[0], up_pad[1], up_pad[2],
                    dn_pad[0], dn_pad[1], dn_pad[2],
                    ev.get("direction"),
                    ev.get("breakout_ref"),
                    ev.get("distance_past_range"),
                    row.get("open"),
                    row.get("high"),
                    row.get("low"),
                    row.get("close"),
                    row.get("volume"),
                    row.get("ema20"),
                    row.get("ema50"),
                    row.get("ema200"),
                    row.get("atr14"),
                    row.get("rsi14"),
                    row.get("vol_sma20"),
                ]
                append_excel_row(EXCEL_LOG_PATH, xl)

    print_daily_summary(stats)
    if WRITE_DAILY_SUMMARY_TO_EXCEL:
        write_daily_summary(EXCEL_LOG_PATH, stats)

def get_test_excel_path() -> str:
    """
    Create a unique Excel filename for TEST_MODE so logs never mix
    with live daytime runs.
    """
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{TEST_LOG_FILE_PREFIX}_{ts}.xlsx"

def run_monitor():
    # If we're running after-hours TEST_MODE, write to a separate file to avoid mixing with daytime logs
    global EXCEL_LOG_PATH
    if TEST_MODE and SEPARATE_TEST_LOGS:
        EXCEL_LOG_PATH = get_test_excel_path()
        log.warning(f"TEST_MODE: writing logs to separate file: {EXCEL_LOG_PATH}")

    ensure_excel_exists(EXCEL_LOG_PATH)
    symbols_cfg, fallback_exchanges = load_symbols_config(SYMBOLS_FILE)

    tv_user, tv_pass = load_tv_credentials_from_keyring()
    if tv_user and tv_pass:
        tv = TvDatafeed(username=tv_user, password=tv_pass)
        log.info("✅ tvDatafeed logged in using OS keychain.")
    else:
        tv = TvDatafeed()
        log.warning("⚠️ tvDatafeed started WITHOUT login. Run setup_keyring.py for best reliability.")

    states: Dict[str, RollingRangeState] = {s["symbol"]: RollingRangeState() for s in symbols_cfg}
    stats: Dict[str, SymbolDailyStats] = {s["symbol"]: SymbolDailyStats(symbol=s["symbol"]) for s in symbols_cfg}

    while True:
        if MARKET_HOURS_ONLY and not in_us_market_hours():
            if EXIT_AT_MARKET_CLOSE:
                log.info("Market is closed. Exiting and generating daily snapshot...")
                print_daily_summary(stats)
                if WRITE_DAILY_SUMMARY_TO_EXCEL:
                    write_daily_summary(EXCEL_LOG_PATH, stats)
                return
            log.info("Market closed/outside configured hours. Sleeping 60s...")
            time.sleep(60)
            continue

        loop_start = time.time()

        for s in symbols_cfg:
            sym = s["symbol"]
            preferred_exch = s.get("exchange")

            df, exch_used = fetch_with_exchange_fallback(
                tv, sym, preferred_exch, fallback_exchanges,
                CANDLE_INTERVAL, HISTORY_BARS
            )
            if df.empty or len(df) < 5:
                continue

            # Bar-close guard
            bar_time = str(df.iloc[-1]["time"])
            if states[sym].last_bar_time_seen == bar_time:
                continue
            states[sym].last_bar_time_seen = bar_time

            events = evaluate_rolling_range(sym, df, states[sym], CANDLE_INTERVAL)
            for ev in events:
                event_type = ev.get("type")

                # Optional logging controls
                if event_type == "RANGE_READY" and not LOG_RANGE_READY_TO_EXCEL:
                    continue
                if event_type == "BREAKOUT" and not LOG_BREAKOUT_TO_EXCEL:
                    continue

                row = ev.get("row", {})

                up = ev.get("up_targets") or []
                dn = ev.get("dn_targets") or []
                up_pad = (up + [None]*3)[:3]
                dn_pad = (dn + [None]*3)[:3]

                from datetime import datetime
                xl = [
                    datetime.now().isoformat(timespec="seconds"),
                    sym,
                    exch_used,
                    str(CANDLE_INTERVAL),
                    ev.get("type"),
                    ev.get("time"),
                    ev.get("range_start"),
                    ev.get("range_end"),
                    ev.get("range_high"),
                    ev.get("range_low"),
                    ev.get("range_width"),
                    ev.get("next_high_target"),
                    ev.get("next_low_target"),
                    ev.get("target_step"),
                    ev.get("direction"),
                    ev.get("breakout_ref"),
                    ev.get("distance_past_range"),
                    row.get("open"),
                    row.get("high"),
                    row.get("low"),
                    row.get("close"),
                    row.get("volume"),
                    row.get("ema20"),
                    row.get("ema50"),
                    row.get("ema200"),
                    row.get("atr14"),
                    row.get("rsi14"),
                    row.get("vol_sma20"),
                    up_pad[0], up_pad[1], up_pad[2],
                    dn_pad[0], dn_pad[1], dn_pad[2],
                ]
                append_excel_row(EXCEL_LOG_PATH, xl)

                if SEND_EMAIL_ALERTS and EMAIL_ON_BREAKOUT and ev.get("type") == "BREAKOUT":
                    subject = f"[ROLLING ORB] {sym} BREAKOUT {ev.get('direction')} ({CANDLE_INTERVAL})"
                    body = f"Logged breakout for {sym} at {ev.get('time')}. Close={row.get('close')}\n"
                    send_email(EMAIL_CFG, subject, body)

                log.info(f"Logged: {sym} {ev.get('type')} @ {ev.get('time')}")

        elapsed = time.time() - loop_start
        time.sleep(max(1, POLL_SECONDS - int(elapsed)))


if __name__ == "__main__":
    run_monitor()
# -------------------------
# Log file naming helpers
# -------------------------
def get_test_excel_path() -> str:
    """Create a unique Excel filename for TEST_MODE so logs never mix with live-day runs."""
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{TEST_LOG_FILE_PREFIX}_{ts}.xlsx"


