\
"""
VCRE Monitor (TradingView candles via tvDatafeed) + Email alerts + Excel log.

What this does
--------------
- Fetches OHLCV candles for each symbol from TradingView via tvDatafeed
- Recomputes your Pine indicator logic (Volume-Confirmed Reversal Engine)
- Alerts only after a candle CLOSES (bar-close guard)
- Emails setup + signal alerts (optional filters)
- Logs every alert into an Excel .xlsx file

Quick start
-----------
1) pip install -r requirements.txt
2) Create/modify symbols.json
3) Store TradingView creds in OS keychain (one-time):
      python setup_keyring.py
4) Fill EMAIL_CFG below (SMTP + app password)
5) Run:
      python monitor_vcre.py
"""

import json
import time
import smtplib
import logging
from dataclasses import dataclass
from email.mime.text import MIMEText
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import keyring
from tvDatafeed import TvDatafeed, Interval

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# =========================================================
# USER SETTINGS (EDIT THESE)
# =========================================================

SYMBOLS_FILE = "symbols.json"

# Candle duration (START: 15-minute)
CANDLE_INTERVAL = Interval.in_15_minute

# Polling frequency. You can poll faster than candle close; bar-close guard prevents duplicates.
POLL_SECONDS = 30

# Must be large enough for EMA200 + ATR warmup (>= ~250). 350 is safe.
HISTORY_BARS = 350

# Alerts
ALERT_ON_SETUPS = True
MIN_STARS_TO_ALERT = 3      # Applies only to BUY/SELL. Set to 4 for only 4+ star signals.
COOLDOWN_SECONDS = 10       # Anti-spam per symbol/event (seconds)

# Excel log file (created automatically if missing)
EXCEL_LOG_PATH = "vcre_alert_log.xlsx"

# Debug logging
DEBUG = True  # Set False for quieter logs

# Market hours mode (US regular session)
MARKET_HOURS_ONLY = True
RUN_EXTENDED_HOURS = False   # True = run weekdays even outside 9:30–16:00 ET
US_MARKET_TZ = "America/New_York"
MARKET_OPEN = (9, 30)
MARKET_CLOSE = (16, 0)

# -------------------------
# Stop-loss preference (you can change this later)
# -------------------------
STOP_MODE = "ANCHOR_ATR_BUFFER"
# Options:
#   "ANCHOR"               -> stop at anchor extreme (structure only)
#   "ATR"                  -> stop at entry ± ATR_STOP_MULT*ATR
#   "ANCHOR_ATR_BUFFER"    -> stop at anchor ± ATR_BUFFER_MULT*ATR (recommended default)
#   "WIDER_OF_ANCHOR_ATR"  -> wider of anchor/ATR (more room, more risk)
#   "TIGHTER_OF_ANCHOR_ATR"-> tighter of anchor/ATR (less risk, more stopouts)

ATR_STOP_MULT = 1.5          # used by ATR modes
ATR_BUFFER_MULT = 0.25       # used by ANCHOR_ATR_BUFFER (tune 0.10–0.50)


# =========================================================
# STRATEGY PARAMETERS (match Pine defaults)
# =========================================================
LOOK_BACK = 20
CONFIRM_IN = 3

USE_VOLUME_FILTER = True
VOL_MA_LENGTH = 20
ANCHOR_VOL_MULTIPLIER = 2.0

USE_CONFIRMATION_VOLUME = True
CONFIRM_VOL_MULTIPLIER = 1.2

MACRO_TREND_LEN = 200
ATR_LEN = 14


# =========================================================
# EMAIL CONFIG (fill these)
# =========================================================
@dataclass
class EmailConfig:
    smtp_host: str
    smtp_port: int
    username: str
    password: str
    from_addr: str
    to_addrs: List[str]
    use_tls: bool = True


EMAIL_CFG = EmailConfig(
    smtp_host="smtp.gmail.com",
    smtp_port=587,
    username="YOUR_EMAIL@gmail.com",
    password="YOUR_APP_PASSWORD",      # Use an App Password for Gmail/Outlook
    from_addr="YOUR_EMAIL@gmail.com",
    to_addrs=["YOUR_EMAIL@gmail.com"],
    use_tls=True
)


# =========================================================
# LOGGING
# =========================================================
logging.basicConfig(
    level=logging.DEBUG if DEBUG else logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
log = logging.getLogger("VCRE")


# =========================================================
# STATE / SETTINGS
# =========================================================
@dataclass
class StrategyConfig:
    look_back: int = LOOK_BACK
    confirm_in: int = CONFIRM_IN
    use_volume_filter: bool = USE_VOLUME_FILTER
    vol_ma_length: int = VOL_MA_LENGTH
    anchor_vol_multiplier: float = ANCHOR_VOL_MULTIPLIER
    use_confirmation_volume: bool = USE_CONFIRMATION_VOLUME
    confirm_vol_multiplier: float = CONFIRM_VOL_MULTIPLIER
    macro_trend_len: int = MACRO_TREND_LEN
    atr_len: int = ATR_LEN


@dataclass
class SymbolState:
    # Bull setup
    bull_setup_active: bool = False
    bull_anchor_low: float = 0.0
    bull_anchor_high: float = 0.0
    bull_bars_since_setup: int = 0
    bull_anchor_has_vol: bool = False
    bull_anchor_time: Optional[str] = None

    # Bear setup
    bear_setup_active: bool = False
    bear_anchor_low: float = 0.0
    bear_anchor_high: float = 0.0
    bear_bars_since_setup: int = 0
    bear_anchor_has_vol: bool = False
    bear_anchor_time: Optional[str] = None

    # Bar-close guard: evaluate once per new closed candle timestamp
    last_bar_time_seen: Optional[str] = None

    # Anti-spam
    last_alert_time: float = 0.0
    last_alert_key: Optional[str] = None


# =========================================================
# MARKET HOURS CHECK
# =========================================================
def in_us_market_hours() -> bool:
    from datetime import datetime, time as dtime
    import zoneinfo

    tz = zoneinfo.ZoneInfo(US_MARKET_TZ)
    now = datetime.now(tz)

    # Weekend
    if now.weekday() >= 5:
        return False

    # Weekday but outside RTH
    if RUN_EXTENDED_HOURS:
        return True

    open_t = dtime(MARKET_OPEN[0], MARKET_OPEN[1])
    close_t = dtime(MARKET_CLOSE[0], MARKET_CLOSE[1])
    return open_t <= now.time() <= close_t


# =========================================================
# EMAIL
# =========================================================
def send_email(cfg: EmailConfig, subject: str, body: str) -> None:
    msg = MIMEText(body, "plain")
    msg["Subject"] = subject
    msg["From"] = cfg.from_addr
    msg["To"] = ", ".join(cfg.to_addrs)

    with smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=30) as server:
        if cfg.use_tls:
            server.starttls()
        server.login(cfg.username, cfg.password)
        server.sendmail(cfg.from_addr, cfg.to_addrs, msg.as_string())


# =========================================================
# INDICATORS (Pine-like)
# =========================================================
def sma(series: pd.Series, length: int) -> pd.Series:
    return series.rolling(length, min_periods=length).mean()

def ema(series: pd.Series, length: int) -> pd.Series:
    return series.ewm(span=length, adjust=False, min_periods=length).mean()

def atr_wilder(df: pd.DataFrame, length: int = 14) -> pd.Series:
    high = df["high"]
    low = df["low"]
    close = df["close"]
    prev_close = close.shift(1)

    tr = pd.concat(
        [(high - low), (high - prev_close).abs(), (low - prev_close).abs()],
        axis=1
    ).max(axis=1)

    # Wilder RMA approximation: ewm(alpha=1/len)
    return tr.ewm(alpha=1/length, adjust=False, min_periods=length).mean()

def crossover(prev_close: float, curr_close: float, level: float) -> bool:
    # Pine ta.crossover(close, level): close[1] <= level and close > level
    return (prev_close <= level) and (curr_close > level)

def crossunder(prev_close: float, curr_close: float, level: float) -> bool:
    # Pine ta.crossunder(close, level): close[1] >= level and close < level
    return (prev_close >= level) and (curr_close < level)

def count_breakout_candles(df: pd.DataFrame, is_bullish: bool, lookback: int) -> int:
    """
    Matches Pine's early-exit loop using the current bar close:
      bullish: close < low[i] for i=1..lookback
      bearish: close > high[i] for i=1..lookback
    """
    c0 = float(df["close"].iloc[-1])
    count = 0
    for i in range(1, lookback + 1):
        if is_bullish:
            if c0 < float(df["low"].iloc[-1 - i]):
                count += 1
            else:
                break
        else:
            if c0 > float(df["high"].iloc[-1 - i]):
                count += 1
            else:
                break
    return count

def calc_signal_strength(cfg: StrategyConfig, is_bullish: bool,
                         anchor_has_vol: bool, confirm_vol_ok: bool,
                         current_close: float, ema_value: float) -> int:
    """
    Base 3 stars +:
      +1 if anchor had high volume
      +1 if confirmation candle had high volume (if enabled)
      +1 if aligned with EMA macro trend
    Capped at 5.
    """
    score = 3
    score += 1 if anchor_has_vol else 0
    score += 1 if (cfg.use_confirmation_volume and confirm_vol_ok) else 0
    score += 1 if (current_close > ema_value if is_bullish else current_close < ema_value) else 0
    return min(score, 5)

def stars(score: int) -> str:
    return {3: "★★★", 4: "★★★★", 5: "★★★★★"}.get(score, "★★★")


# =========================================================
# SYMBOL CONFIG + FETCH (exchange fallback)
# =========================================================
def load_symbols_config(path: str) -> Tuple[List[dict], List[str]]:
    """
    Reads symbols.json. Example entry:
      { "symbol": "TSLA", "exchange": "NASDAQ", "enabled": true }
    """
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    fallback = cfg.get("default_exchange_fallback", ["AMEX", "NASDAQ", "NYSE"])
    symbols = [s for s in cfg.get("symbols", []) if s.get("enabled", True)]
    return symbols, fallback

def normalize_tv_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    tvDatafeed returns a DataFrame indexed by datetime with columns:
      open, high, low, close, volume
    We normalize to: time, open, high, low, close, volume
    """
    df = df.copy().reset_index()
    if "datetime" in df.columns:
        df.rename(columns={"datetime": "time"}, inplace=True)
    elif "date" in df.columns:
        df.rename(columns={"date": "time"}, inplace=True)
    else:
        df.rename(columns={df.columns[0]: "time"}, inplace=True)

    needed = ["time", "open", "high", "low", "close", "volume"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    return df[needed].sort_values("time").reset_index(drop=True)

def fetch_with_exchange_fallback(
    tv: TvDatafeed,
    symbol: str,
    preferred_exchange: Optional[str],
    fallback_exchanges: List[str],
    interval: Interval,
    n_bars: int,
    retries_per_exchange: int = 2
) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Tries preferred exchange first, then fallback list (e.g. AMEX -> NASDAQ -> NYSE).
    Returns (df, exchange_used) or (empty, None).
    """
    exchanges_to_try: List[str] = []
    if preferred_exchange:
        exchanges_to_try.append(preferred_exchange)
    for ex in fallback_exchanges:
        if ex not in exchanges_to_try:
            exchanges_to_try.append(ex)

    last_err = None
    for exch in exchanges_to_try:
        for attempt in range(1, retries_per_exchange + 1):
            try:
                raw = tv.get_hist(symbol=symbol, exchange=exch, interval=interval, n_bars=n_bars)
                if raw is None or raw.empty:
                    raise ValueError("Empty dataframe")
                return normalize_tv_df(raw), exch
            except Exception as e:
                last_err = e
                time.sleep(0.5 * attempt)

    log.warning(f"[{symbol}] no data from exchanges {exchanges_to_try}. last_err={last_err}")
    return pd.DataFrame(), None


# =========================================================
# STOP / TARGETS (trade plan)
# =========================================================
def build_trade_plan(event: dict, atr_value: float) -> dict:
    """
    Stop/target suggestions derived from indicator structure + ATR.

    How to change stop behavior:
      - Set STOP_MODE near the top of this file.
      - Tune ATR_STOP_MULT and ATR_BUFFER_MULT.

    STOP_MODE choices:
      - ANCHOR: stop at anchor extreme
      - ATR: stop at entry ± ATR_STOP_MULT*ATR
      - ANCHOR_ATR_BUFFER: stop at anchor ± ATR_BUFFER_MULT*ATR  (recommended default)
      - WIDER_OF_ANCHOR_ATR: uses wider stop of anchor/ATR
      - TIGHTER_OF_ANCHOR_ATR: uses tighter stop of anchor/ATR
    """
    etype = event["type"]
    close = float(event["close"])  # alert candle close
    anchor_high = float(event.get("anchor_high", np.nan))
    anchor_low = float(event.get("anchor_low", np.nan))

    plan = {
        "entry_idea": None,
        "stop_anchor": None,
        "stop_atr": None,
        "primary_stop": None,
        "stop_mode_used": STOP_MODE,
        "risk_per_share": None,
        "target_1R": None,
        "target_2R": None,
    }

    if np.isnan(anchor_high) or np.isnan(anchor_low) or np.isnan(atr_value):
        return plan

    if etype in ("BULL_SETUP", "BUY"):
        plan["entry_idea"] = (
            f"Conservative: Buy-stop just above anchor high ({anchor_high:.2f}). "
            f"Aggressive: enter near close/next open (~{close:.2f})."
        )
        stop_anchor = anchor_low
        stop_atr = close - (ATR_STOP_MULT * atr_value)
        stop_anchor_atr_buffer = anchor_low - (ATR_BUFFER_MULT * atr_value)

        if STOP_MODE == "ANCHOR":
            primary = stop_anchor
        elif STOP_MODE == "ATR":
            primary = stop_atr
        elif STOP_MODE == "ANCHOR_ATR_BUFFER":
            primary = stop_anchor_atr_buffer
        elif STOP_MODE == "WIDER_OF_ANCHOR_ATR":
            primary = min(stop_anchor, stop_atr)  # lower stop = wider (long)
        elif STOP_MODE == "TIGHTER_OF_ANCHOR_ATR":
            primary = max(stop_anchor, stop_atr)  # higher stop = tighter (long)
        else:
            primary = stop_anchor_atr_buffer

        plan["stop_anchor"] = stop_anchor
        plan["stop_atr"] = stop_atr
        plan["primary_stop"] = primary

        risk = close - primary
        if risk > 0:
            plan["risk_per_share"] = risk
            plan["target_1R"] = close + risk
            plan["target_2R"] = close + 2 * risk

    elif etype in ("BEAR_SETUP", "SELL"):
        plan["entry_idea"] = (
            f"Conservative: Sell-stop just below anchor low ({anchor_low:.2f}). "
            f"Aggressive: enter near close/next open (~{close:.2f})."
        )
        stop_anchor = anchor_high
        stop_atr = close + (ATR_STOP_MULT * atr_value)
        stop_anchor_atr_buffer = anchor_high + (ATR_BUFFER_MULT * atr_value)

        if STOP_MODE == "ANCHOR":
            primary = stop_anchor
        elif STOP_MODE == "ATR":
            primary = stop_atr
        elif STOP_MODE == "ANCHOR_ATR_BUFFER":
            primary = stop_anchor_atr_buffer
        elif STOP_MODE == "WIDER_OF_ANCHOR_ATR":
            primary = max(stop_anchor, stop_atr)  # higher stop = wider (short)
        elif STOP_MODE == "TIGHTER_OF_ANCHOR_ATR":
            primary = min(stop_anchor, stop_atr)  # lower stop = tighter (short)
        else:
            primary = stop_anchor_atr_buffer

        plan["stop_anchor"] = stop_anchor
        plan["stop_atr"] = stop_atr
        plan["primary_stop"] = primary

        risk = primary - close
        if risk > 0:
            plan["risk_per_share"] = risk
            plan["target_1R"] = close - risk
            plan["target_2R"] = close - 2 * risk

    return plan


# =========================================================
# EXCEL LOGGING
# =========================================================
EXCEL_HEADERS = [
    "log_time_local",
    "symbol",
    "exchange_used",
    "interval",
    "event_type",
    "score",
    "stars",
    "bar_time",
    "close",
    "anchor_time",
    "anchor_high",
    "anchor_low",
    "volume",
    "avg_volume",
    "confirm_vol_ok",
    "ema_macro",
    "atr",
    "stop_mode",
    "entry_idea",
    "stop_anchor",
    "stop_atr",
    "primary_stop",
    "risk_per_share",
    "target_1R",
    "target_2R",
]

def ensure_excel_exists(path: str) -> None:
    try:
        load_workbook(path)
        return
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "alerts"
        ws.append(EXCEL_HEADERS)

        for i, h in enumerate(EXCEL_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 14), 45)

        wb.save(path)

def append_excel_row(path: str, row: List) -> None:
    wb = load_workbook(path)
    ws = wb["alerts"]
    ws.append(row)
    wb.save(path)


# =========================================================
# CORE EVALUATION (Pine port)
# =========================================================
def evaluate_symbol(symbol: str, df: pd.DataFrame, cfg: StrategyConfig, state: SymbolState) -> List[dict]:
    """
    Debug points:
      - After indicator columns computed (df.tail(2))
      - After anchor detection (bull_count/bear_count, volume flags)
      - Signal conditions (crossover/crossunder)
    """
    events: List[dict] = []

    min_hist = max(cfg.macro_trend_len, cfg.vol_ma_length, cfg.atr_len, cfg.look_back + 2)
    if len(df) < min_hist:
        return events

    df = df.copy()
    df["avg_volume"] = sma(df["volume"], cfg.vol_ma_length)
    df["ema_macro"] = ema(df["close"], cfg.macro_trend_len)
    df["atr"] = atr_wilder(df, cfg.atr_len)

    if DEBUG:
        log.debug(f"[{symbol}] df.tail(2):\n{df.tail(2)}")

    curr = df.iloc[-1]
    prev = df.iloc[-2]

    time_str = str(curr["time"])
    c0 = float(curr["close"])
    c1 = float(prev["close"])
    v0 = float(curr["volume"])

    avg_v = float(curr["avg_volume"]) if not np.isnan(curr["avg_volume"]) else np.nan
    ema_macro = float(curr["ema_macro"]) if not np.isnan(curr["ema_macro"]) else np.nan
    atr_val = float(curr["atr"]) if not np.isnan(curr["atr"]) else np.nan

    is_anchor_vol_ok = False
    is_confirm_vol_ok = False
    if not np.isnan(avg_v):
        is_anchor_vol_ok = v0 > avg_v * cfg.anchor_vol_multiplier
        is_confirm_vol_ok = v0 > avg_v * cfg.confirm_vol_multiplier

    bull_count = count_breakout_candles(df, True, cfg.look_back)
    bear_count = count_breakout_candles(df, False, cfg.look_back)
    is_bull_anchor_price = (bull_count == cfg.look_back)
    is_bear_anchor_price = (bear_count == cfg.look_back)

    if DEBUG:
        log.debug(
            f"[{symbol}] bull_count={bull_count}, bear_count={bear_count}, "
            f"is_anchor_vol_ok={is_anchor_vol_ok}, avg_v={avg_v}, v0={v0}"
        )

    # STEP 1: Anchor detection
    if is_bull_anchor_price and (is_anchor_vol_ok or not cfg.use_volume_filter):
        state.bear_setup_active = False
        state.bull_setup_active = True
        state.bull_anchor_low = float(curr["low"])
        state.bull_anchor_high = float(curr["high"])
        state.bull_bars_since_setup = 0
        state.bull_anchor_has_vol = bool(is_anchor_vol_ok)
        state.bull_anchor_time = time_str

        events.append({
            "type": "BULL_SETUP",
            "time": time_str,
            "close": c0,
            "volume": v0,
            "avg_volume": avg_v,
            "anchor_high": state.bull_anchor_high,
            "anchor_low": state.bull_anchor_low,
            "anchor_has_vol": state.bull_anchor_has_vol,
            "ema_macro": ema_macro,
            "atr": atr_val,
            "confirm_vol_ok": None,
            "score": None,
            "stars": None,
            "anchor_time": state.bull_anchor_time
        })

    if is_bear_anchor_price and (is_anchor_vol_ok or not cfg.use_volume_filter):
        state.bull_setup_active = False
        state.bear_setup_active = True
        state.bear_anchor_low = float(curr["low"])
        state.bear_anchor_high = float(curr["high"])
        state.bear_bars_since_setup = 0
        state.bear_anchor_has_vol = bool(is_anchor_vol_ok)
        state.bear_anchor_time = time_str

        events.append({
            "type": "BEAR_SETUP",
            "time": time_str,
            "close": c0,
            "volume": v0,
            "avg_volume": avg_v,
            "anchor_high": state.bear_anchor_high,
            "anchor_low": state.bear_anchor_low,
            "anchor_has_vol": state.bear_anchor_has_vol,
            "ema_macro": ema_macro,
            "atr": atr_val,
            "confirm_vol_ok": None,
            "score": None,
            "stars": None,
            "anchor_time": state.bear_anchor_time
        })

    # STEP 2: Setup management & invalidation
    if state.bull_setup_active and (not is_bull_anchor_price):
        state.bull_bars_since_setup += 1
        timeout = state.bull_bars_since_setup > cfg.confirm_in
        new_low = c0 < state.bull_anchor_low
        if timeout or new_low:
            state.bull_setup_active = False

    if state.bear_setup_active and (not is_bear_anchor_price):
        state.bear_bars_since_setup += 1
        timeout = state.bear_bars_since_setup > cfg.confirm_in
        new_high = c0 > state.bear_anchor_high
        if timeout or new_high:
            state.bear_setup_active = False

    # STEP 3: Confirmation & signals
    if state.bull_setup_active and state.bull_bars_since_setup > 0:
        if crossover(c1, c0, state.bull_anchor_high):
            score = calc_signal_strength(cfg, True, state.bull_anchor_has_vol, is_confirm_vol_ok, c0, ema_macro)
            events.append({
                "type": "BUY",
                "time": time_str,
                "close": c0,
                "volume": v0,
                "avg_volume": avg_v,
                "anchor_time": state.bull_anchor_time,
                "anchor_high": state.bull_anchor_high,
                "anchor_low": state.bull_anchor_low,
                "score": score,
                "stars": stars(score),
                "ema_macro": ema_macro,
                "confirm_vol_ok": bool(is_confirm_vol_ok),
                "atr": atr_val
            })
            state.bull_setup_active = False

    if state.bear_setup_active and state.bear_bars_since_setup > 0:
        if crossunder(c1, c0, state.bear_anchor_low):
            score = calc_signal_strength(cfg, False, state.bear_anchor_has_vol, is_confirm_vol_ok, c0, ema_macro)
            events.append({
                "type": "SELL",
                "time": time_str,
                "close": c0,
                "volume": v0,
                "avg_volume": avg_v,
                "anchor_time": state.bear_anchor_time,
                "anchor_high": state.bear_anchor_high,
                "anchor_low": state.bear_anchor_low,
                "score": score,
                "stars": stars(score),
                "ema_macro": ema_macro,
                "confirm_vol_ok": bool(is_confirm_vol_ok),
                "atr": atr_val
            })
            state.bear_setup_active = False

    return events


# =========================================================
# ALERT CONTROL
# =========================================================
def should_send(state: SymbolState, key: str, cooldown: int) -> bool:
    now = time.time()
    if state.last_alert_key == key:
        return False
    if (now - state.last_alert_time) < cooldown:
        return False
    state.last_alert_time = now
    state.last_alert_key = key
    return True


# =========================================================
# TV CREDENTIALS FROM KEYRING
# =========================================================
def load_tv_credentials_from_keyring() -> Tuple[Optional[str], Optional[str]]:
    user = keyring.get_password("tradingview", "username")
    pwd = keyring.get_password("tradingview", "password")
    return user, pwd


# =========================================================
# MAIN LOOP
# =========================================================
def run_monitor():
    ensure_excel_exists(EXCEL_LOG_PATH)
    symbols_cfg, fallback_exchanges = load_symbols_config(SYMBOLS_FILE)

    # tvDatafeed uses username/password (Playwright tv_state.json is not used here)
    tv_user, tv_pass = load_tv_credentials_from_keyring()
    if tv_user and tv_pass:
        tv = TvDatafeed(username=tv_user, password=tv_pass)
        log.info("✅ TvDatafeed logged in using OS keychain (keyring).")
    else:
        tv = TvDatafeed()
        log.warning("⚠️ TvDatafeed started WITHOUT login. Run setup_keyring.py for best reliability.")

    cfg = StrategyConfig()
    states: Dict[str, SymbolState] = {s["symbol"]: SymbolState() for s in symbols_cfg}

    while True:
        # Market-hours-only throttle
        if MARKET_HOURS_ONLY and not in_us_market_hours():
            log.info("Market closed/outside configured hours. Sleeping 60s...")
            time.sleep(60)
            continue

        loop_start = time.time()

        for s in symbols_cfg:
            sym = s["symbol"]
            preferred_exch = s.get("exchange")

            if DEBUG:
                log.debug(f"--- Processing {sym} (preferred_exchange={preferred_exch}) ---")

            # 1) Fetch data (exchange fallback included)
            df, exch_used = fetch_with_exchange_fallback(
                tv, sym, preferred_exch, fallback_exchanges,
                CANDLE_INTERVAL, HISTORY_BARS
            )
            if df.empty:
                continue

            # 2) Bar-close guard: only evaluate on NEW candle close
            bar_time = str(df.iloc[-1]["time"])
            if states[sym].last_bar_time_seen == bar_time:
                continue
            states[sym].last_bar_time_seen = bar_time

            # 3) Evaluate indicator logic for this new closed bar
            events = evaluate_symbol(sym, df, cfg, states[sym])
            if not events:
                continue

            # 4) For each event: build trade plan, send email, log to Excel
            for ev in events:
                etype = ev["type"]

                if etype in ("BULL_SETUP", "BEAR_SETUP") and not ALERT_ON_SETUPS:
                    continue

                if etype in ("BUY", "SELL"):
                    score = int(ev["score"])
                    if score < MIN_STARS_TO_ALERT:
                        continue

                key = f"{sym}:{bar_time}:{etype}:{ev.get('score')}"
                if not should_send(states[sym], key, COOLDOWN_SECONDS):
                    continue

                plan = build_trade_plan(ev, float(ev.get("atr", np.nan)))

                # Email subject
                subject = f"[VCRE] {sym} {etype}"
                if etype in ("BUY", "SELL"):
                    subject += f" {ev.get('stars','')} ({CANDLE_INTERVAL})"
                else:
                    subject += f" ({CANDLE_INTERVAL})"

                # Email body (includes alert candle close)
                body = (
                    f"Symbol: {sym} ({exch_used})\n"
                    f"Interval: {CANDLE_INTERVAL}\n"
                    f"Alert Candle Time: {ev['time']}\n"
                    f"Alert Candle Close: {ev['close']}\n\n"
                    f"Event: {etype}\n"
                )

                if etype in ("BUY", "SELL"):
                    body += (
                        f"Score: {ev['score']} {ev['stars']}\n"
                        f"Anchor Time: {ev.get('anchor_time')}\n"
                        f"Anchor High: {ev.get('anchor_high')}\n"
                        f"Anchor Low: {ev.get('anchor_low')}\n"
                        f"Confirm Volume OK: {ev.get('confirm_vol_ok')}\n"
                        f"Macro EMA({MACRO_TREND_LEN}): {ev.get('ema_macro')}\n"
                        f"ATR({ATR_LEN}): {ev.get('atr')}\n\n"
                    )
                else:
                    body += (
                        f"Anchor High: {ev.get('anchor_high')}\n"
                        f"Anchor Low: {ev.get('anchor_low')}\n"
                        f"Anchor Volume OK: {ev.get('anchor_has_vol')}\n"
                        f"Macro EMA({MACRO_TREND_LEN}): {ev.get('ema_macro')}\n"
                        f"ATR({ATR_LEN}): {ev.get('atr')}\n\n"
                    )

                body += (
                    "Trade Plan Ideas (mechanical):\n"
                    f"- Entry: {plan.get('entry_idea')}\n"
                    f"- Stop mode: {plan.get('stop_mode_used')}\n"
                    f"- Stop (anchor): {plan.get('stop_anchor')}\n"
                    f"- Stop (ATR): {plan.get('stop_atr')}\n"
                    f"- Primary stop: {plan.get('primary_stop')}\n"
                    f"- Risk/share: {plan.get('risk_per_share')}\n"
                    f"- Target 1R: {plan.get('target_1R')}\n"
                    f"- Target 2R: {plan.get('target_2R')}\n\n"
                    "Note: Derived from indicator structure + ATR; not personalized advice.\n"
                )

                send_email(EMAIL_CFG, subject, body)

                # Excel log row
                from datetime import datetime
                row = [
                    datetime.now().isoformat(timespec="seconds"),
                    sym,
                    exch_used,
                    str(CANDLE_INTERVAL),
                    etype,
                    ev.get("score"),
                    ev.get("stars"),
                    ev.get("time"),
                    ev.get("close"),
                    ev.get("anchor_time"),
                    ev.get("anchor_high"),
                    ev.get("anchor_low"),
                    ev.get("volume"),
                    ev.get("avg_volume"),
                    ev.get("confirm_vol_ok"),
                    ev.get("ema_macro"),
                    ev.get("atr"),
                    plan.get("stop_mode_used"),
                    plan.get("entry_idea"),
                    plan.get("stop_anchor"),
                    plan.get("stop_atr"),
                    plan.get("primary_stop"),
                    plan.get("risk_per_share"),
                    plan.get("target_1R"),
                    plan.get("target_2R"),
                ]
                append_excel_row(EXCEL_LOG_PATH, row)

                log.info(f"Alert sent + logged: {sym} {etype} @ {ev['time']} close={ev['close']}")

        elapsed = time.time() - loop_start
        time.sleep(max(1, POLL_SECONDS - int(elapsed)))


if __name__ == "__main__":
    run_monitor()
